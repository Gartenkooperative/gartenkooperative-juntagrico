import urllib

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import make_naive
from django.utils.translation import gettext_lazy as _
from juntagrico.config import Config
from juntagrico.entity.jobs import ActivityArea
from juntagrico.util.temporal import start_of_business_year, end_of_business_year
from openpyxl import Workbook

from .forms import DateRangeForm
from .utils import TemporalData, assignments_by, slots_by, assignments_by_subscription, members_with_assignments, date_from_get


@staff_member_required
def assignments(request, trunc='week'):
    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    temporal_data = TemporalData(trunc, start_date, end_date)
    data = temporal_data.data_to_dict(assignments_by)
    done_jobs = list(data.values())
    available_slots = list(temporal_data.data_to_dict(slots_by, 'available').values())

    renderdict = {
        'trunc_name': temporal_data.trunc_adjective().capitalize() + 'e',
        'start_date': start_date,
        'end_date': end_date,
        'labels': [temporal_data.date_to_label(date) for date in data.keys()],
        'done_jobs': done_jobs,
        'available_slots': available_slots,
        'mobilization': [i / j if j > 0 else 0 for i, j in zip(done_jobs, available_slots)],
        'query': urllib.parse.urlencode(request.GET),
        'date_form': DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})
    }
    return render(request, 'jst/assignments.html', renderdict)


@staff_member_required
def assignments_export(request):
    activity_area = ActivityArea.objects.filter(pk=request.GET.get('activity_area', None)).first()

    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    filename = '{}_{}_{}assignments.xlsx'.format(
        start_date.strftime('%Y-%m-%d'),
        end_date.strftime('%Y-%m-%d'),
        str(activity_area.pk) + '_' if activity_area else ''
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = Workbook()

    # Sheet 1: assignments by subscription
    ws1 = wb.active
    ws1.title = "assignments by subscription"

    # header
    ws1.cell(1, 1, u"{}".format(Config.vocabulary('member_pl')))
    ws1.column_dimensions['A'].width = 40
    ws1.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws1.column_dimensions['B'].width = 17
    ws1.cell(1, 3, u"{}".format(_('{}-Grösse').format(Config.vocabulary('subscription'))))
    ws1.column_dimensions['C'].width = 17

    # data
    for row, subscription in enumerate(assignments_by_subscription(start_date, end_date, activity_area), 2):
        ws1.cell(row, 1, ", ".join([member.get_name() for member in subscription['subscription'].recipients]))
        ws1.cell(row, 2, subscription['assignments'])
        ws1.cell(row, 3, subscription['subscription'].totalsize)

    # Sheet 2: assignments per day
    ws2 = wb.create_sheet(title="assignments per day")

    # header
    ws2.cell(1, 1, u"{}".format(_('Datum')))
    ws2.column_dimensions['A'].width = 20
    ws2.cell(1, 2, u"{}".format(_('Arbeitseinsätze geleistet')))
    ws2.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(assignments_by('day', start_date, end_date, activity_area), 2):
        ws2.cell(row, 1, make_naive(assignment['day']))
        ws2.cell(row, 2, assignment['count'])

    # Sheet 3: slots by day
    ws3 = wb.create_sheet(title="slots per day")

    # header
    ws3.cell(1, 1, u"{}".format(_('Datum')))
    ws3.column_dimensions['A'].width = 20
    ws3.cell(1, 2, u"{}".format(_('Arbeitseinsätze ausgeschrieben')))
    ws3.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(slots_by('day', start_date, end_date, activity_area), 2):
        ws3.cell(row, 1, make_naive(assignment['day']))
        ws3.cell(row, 2, assignment['available'])

    # Sheet 4: assignments per member
    ws4 = wb.create_sheet(title="assignments per member")

    # header
    ws4.cell(1, 1, u"{}".format(Config.vocabulary('member')))
    ws4.column_dimensions['A'].width = 40
    ws4.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws4.column_dimensions['B'].width = 17

    # data
    members = members_with_assignments(start_date, end_date, activity_area)
    for row, member in enumerate(members, 2):
        ws4.cell(row, 1, u"{}".format(member))
        ws4.cell(row, 2, member.assignments)

    wb.save(response)
    return response
